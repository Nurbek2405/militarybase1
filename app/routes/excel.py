from flask import render_template, request, redirect, url_for, flash, send_file
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import xlwt
import xlrd
from datetime import datetime
from models import db, Employee, Examination, PREFLIGHT_CONDITIONS, EXAM_TYPES  # Импорт из models.py
from app.models import recalculate_all_employees

def export_excel_xlsx():
    try:
        employees = Employee.query.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"

        headers = [
            'ID', 'ФИО', 'Дата рождения', 'Должность', 'По приказу № 721', 'Состояние', 'Примечание',
            'ВЛК дата', 'ВЛК диагноз',
            'КМО дата', 'КМО диагноз',
            'УМО дата', 'УМО диагноз',
            'КМО2 дата', 'КМО2 диагноз'
        ]
        ws.append(headers)

        for emp in employees:
            vlk_date = vlk_diagnosis = kmo_date = kmo_diagnosis = umo_date = umo_diagnosis = kmo2_date = kmo2_diagnosis = ''
            for exam in emp.examinations:
                if exam.exam_type == 'ВЛК':
                    vlk_date = exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else ''
                    vlk_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'КМО':
                    kmo_date = exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else ''
                    kmo_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'УМО':
                    umo_date = exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else ''
                    umo_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'КМО2':
                    kmo2_date = exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else ''
                    kmo2_diagnosis = exam.diagnosis or ''

            row = [
                emp.id, emp.fio,
                emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '',
                emp.position, emp.order_no, emp.preflight_condition, emp.note or '',
                vlk_date, vlk_diagnosis,
                kmo_date, kmo_diagnosis,
                umo_date, umo_diagnosis,
                kmo2_date, kmo2_diagnosis
            ]
            ws.append(row)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='employees.xlsx'
        )
    except Exception as e:
        flash(f"Ошибка экспорта в .xlsx: {str(e)}", 'danger')
        return redirect(url_for('index'))

def export_excel_xls():
    try:
        employees = Employee.query.all()
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Сотрудники')

        headers = [
            'ID', 'ФИО', 'Дата рождения', 'Должность', 'По приказу № 721', 'Состояние', 'Примечание',
            'ВЛК дата', 'ВЛК диагноз',
            'КМО дата', 'КМО диагноз',
            'УМО дата', 'УМО диагноз',
            'КМО2 дата', 'КМО2 диагноз'
        ]
        for col, header in enumerate(headers):
            ws.write(0, col, header)

        row_num = 1
        for emp in employees:
            vlk_date = vlk_diagnosis = kmo_date = kmo_diagnosis = umo_date = umo_diagnosis = kmo2_date = kmo2_diagnosis = ''
            for exam in emp.examinations:
                if exam.exam_type == 'ВЛК':
                    vlk_date = exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else ''
                    vlk_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'КМО':
                    kmo_date = exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else ''
                    kmo_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'УМО':
                    umo_date = exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else ''
                    umo_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'КМО2':
                    kmo2_date = exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else ''
                    kmo2_diagnosis = exam.diagnosis or ''

            row = [
                emp.id, emp.fio,
                emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '',
                emp.position, emp.order_no, emp.preflight_condition, emp.note or '',
                vlk_date, vlk_diagnosis,
                kmo_date, kmo_diagnosis,
                umo_date, umo_diagnosis,
                kmo2_date, kmo2_diagnosis
            ]
            for col, value in enumerate(row):
                ws.write(row_num, col, value)
            row_num += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.ms-excel',
            as_attachment=True,
            download_name='employees.xls'
        )
    except Exception as e:
        flash(f"Ошибка экспорта в .xls: {str(e)}", 'danger')
        return redirect(url_for('index'))

def import_excel():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не выбран!', 'danger')
            return redirect(url_for('import_excel'))
        file = request.files['file']
        if file.filename == '':
            flash('Файл не выбран!', 'danger')
            return redirect(url_for('import_excel'))
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            flash('Поддерживаются только файлы .xlsx и .xls!', 'danger')
            return redirect(url_for('import_excel'))

        try:
            if file.filename.endswith('.xlsx'):
                wb = load_workbook(file)
                ws = wb.active
                rows = list(ws.rows)
                headers = [cell.value for cell in rows[0]]
                data_rows = rows[1:]
            else:  # .xls
                wb = xlrd.open_workbook(file_contents=file.read())
                ws = wb.sheet_by_index(0)
                headers = ws.row_values(0)
                data_rows = [ws.row_values(i) for i in range(1, ws.nrows)]

            skipped_count = 0
            for row in data_rows:
                # Преобразуем данные в зависимости от формата файла
                if file.filename.endswith('.xlsx'):
                    data = [cell.value if cell.value is not None else '' for cell in row]
                else:
                    data = [str(cell) if cell else '' for cell in row]

                # Проверяем минимальную длину строки
                if len(data) < 7:  # ID, ФИО, дата рождения, должность, приказ, состояние, примечание
                    skipped_count += 1
                    continue

                fio = data[1]
                existing_employee = Employee.query.filter_by(fio=fio).first()
                if existing_employee:
                    skipped_count += 1
                    continue

                birth_date_str = data[2]
                if birth_date_str:
                    try:
                        birth_date = datetime.strptime(str(birth_date_str), '%Y-%m-%d').date()
                    except ValueError:
                        flash(f"Неверный формат даты рождения для {fio}: {birth_date_str}", 'danger')
                        continue
                else:
                    birth_date = None

                employee = Employee(
                    fio=fio,
                    birth_date=birth_date,
                    position=data[3],
                    order_no=data[4],
                    preflight_condition=data[5] if data[5] in PREFLIGHT_CONDITIONS else 'Допущен',
                    note=data[6] if data[6] else None
                )
                db.session.add(employee)
                db.session.flush()

                # Импорт осмотров
                if len(data) >= 15:  # Проверяем наличие всех полей осмотров
                    for i, exam_type in enumerate(EXAM_TYPES):
                        date_idx = 7 + i * 2  # ВЛК дата, КМО дата, УМО дата, КМО2 дата
                        diag_idx = 8 + i * 2  # ВЛК диагноз, КМО диагноз, УМО диагноз, КМО2 диагноз
                        exam_date_str = data[date_idx] if date_idx < len(data) else ''
                        diagnosis = data[diag_idx] if diag_idx < len(data) else ''
                        if exam_date_str:
                            try:
                                exam_date = datetime.strptime(str(exam_date_str), '%Y-%m-%d').date()
                                exam = Examination(
                                    employee_id=employee.id,
                                    exam_type=exam_type,
                                    exam_date=exam_date,
                                    diagnosis=diagnosis if diagnosis else None,
                                    note=None
                                )
                                db.session.add(exam)
                            except ValueError:
                                flash(f"Неверный формат даты осмотра {exam_type} для {fio}: {exam_date_str}", 'warning')

            db.session.commit()
            recalculate_all_employees(db.session)
            flash(f'Сотрудники успешно импортированы из Excel! Пропущено дубликатов или некорректных строк: {skipped_count}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f"Ошибка импорта Excel: {str(e)}", 'danger')
            print(f"Ошибка в import_excel(): {str(e)}")
        return redirect(url_for('index'))

    return render_template('import_excel.html')