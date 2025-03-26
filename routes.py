from flask import render_template, request, redirect, url_for, flash, send_file
from models import process_employee_form, calculate_expiry
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import xlwt
import xlrd

def register_routes(app):
    from app import db, Employee, Examination, PREFLIGHT_CONDITIONS

    @app.route('/')
    def index():
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        sort = request.args.get('sort', 'deadline_asc')
        search = request.args.get('search', '')

        query = Employee.query
        if search:
            query = query.filter(Employee.fio.ilike(f'%{search}%'))

        if sort == 'fio_asc':
            query = query.order_by(Employee.fio.asc())
        elif sort == 'fio_desc':
            query = query.order_by(Employee.fio.desc())
        elif sort == 'suspended':
            query = query.filter(Employee.preflight_condition == 'Отстранен')
        elif sort == 'deadline_asc':
            pass

        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        employees = pagination.items
        total_pages = pagination.pages

        if not employees and not Employee.query.filter_by(fio="Тестовый Сотрудник").first():
            test_employee = Employee(
                fio="Тестовый Сотрудник",
                birth_date=datetime.now().date() - timedelta(days=365 * 30),
                position="Пилот",
                order_no="721",
                preflight_condition="Допущен",
                note="Тест подсветки"
            )
            db.session.add(test_employee)
            db.session.commit()
            for exam in [
                {'exam_type': 'ВЛК', 'exam_date': datetime.now().date() - timedelta(days=365), 'diagnosis': 'Годен'},
                {'exam_type': 'КМО', 'exam_date': datetime.now().date() - timedelta(days=85), 'diagnosis': 'Годен'},
                {'exam_type': 'УМО', 'exam_date': datetime.now().date() - timedelta(days=60), 'diagnosis': 'Годен'}
            ]:
                db.session.add(Examination(employee_id=test_employee.id, **exam))
            db.session.commit()
            employees = [test_employee]

        employees_with_expiry = []
        for emp in employees:
            db.session.refresh(emp)  # Синхронизируем объект с базой
            employees_with_expiry.append(calculate_expiry(emp))

        for emp in employees_with_expiry:
            db.session.add(emp['employee'])
        db.session.commit()

        if sort == 'deadline_asc':
            employees_with_expiry.sort(
                key=lambda x: x['min_days_left'] if x['min_days_left'] is not None else float('inf'),
                reverse=True)

        return render_template(
            'index.html',
            employees_with_expiry=employees_with_expiry,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            datetime=datetime,
            preflight_conditions=PREFLIGHT_CONDITIONS
        )

    @app.route('/add', methods=['GET', 'POST'])
    def add():
        if request.method == 'POST':
            try:
                employee_data, examinations = process_employee_form(request.form)
                if not all([employee_data['fio'], employee_data['birth_date'], employee_data['position'],
                            employee_data['order_no']]):
                    flash('Поля ФИО, дата рождения, должность и "По приказу № 721" обязательны!', 'danger')
                    return redirect(url_for('add'))

                existing_employee = Employee.query.filter_by(fio=employee_data['fio']).first()
                if existing_employee:
                    flash(f'Сотрудник с ФИО "{employee_data["fio"]}" уже существует!', 'danger')
                    return redirect(url_for('add'))

                employee = Employee(**employee_data)
                db.session.add(employee)
                db.session.flush()  # Получаем ID сотрудника

                for exam in examinations:
                    db.session.add(Examination(employee_id=employee.id, **exam))

                expiry_data = calculate_expiry(employee)
                employee.preflight_condition = expiry_data['employee'].preflight_condition

                db.session.commit()
                flash('Сотрудник успешно добавлен!', 'success')
                return redirect(url_for('index'))
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка добавления сотрудника: {str(e)}", 'danger')
                return redirect(url_for('add'))

        return render_template('add.html', preflight_conditions=PREFLIGHT_CONDITIONS)

    @app.route('/edit/<int:id>', methods=['GET', 'POST'])
    def edit(id):
        employee = Employee.query.get_or_404(id)
        if request.method == 'POST':
            try:
                employee_data, examinations = process_employee_form(request.form)
                if not all([employee_data['fio'], employee_data['birth_date'], employee_data['position'],
                            employee_data['order_no']]):
                    flash('Поля ФИО, дата рождения, должность и "По приказу № 721" обязательны!', 'danger')
                    return redirect(url_for('edit', id=id))

                existing_employee = Employee.query.filter(Employee.fio == employee_data['fio'],
                                                          Employee.id != id).first()
                if existing_employee:
                    flash(f'Сотрудник с ФИО "{employee_data["fio"]}" уже существует!', 'danger')
                    return redirect(url_for('edit', id=id))

                # Обновляем данные сотрудника
                for key, value in employee_data.items():
                    setattr(employee, key, value)
                Examination.query.filter_by(employee_id=id).delete()  # Удаляем старые осмотры
                for exam in examinations:
                    db.session.add(Examination(employee_id=id, **exam))

                # Пересчитываем сроки и состояние
                expiry_data = calculate_expiry(employee)
                employee.preflight_condition = expiry_data['employee'].preflight_condition

                db.session.commit()
                flash('Данные сотрудника успешно обновлены!', 'success')
                return redirect(url_for('index'))
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка обновления данных: {str(e)}", 'danger')
                return redirect(url_for('edit', id=id))

        examinations = Examination.query.filter_by(employee_id=id).all()
        return render_template('edit.html', employee=employee, examinations=examinations,
                               preflight_conditions=PREFLIGHT_CONDITIONS)

    @app.route('/history/<int:id>')
    def history(id):
        employee = Employee.query.get_or_404(id)
        examinations = Examination.query.filter_by(employee_id=id).order_by(Examination.exam_date.asc()).all()
        return render_template('history.html', employee=employee, examinations=examinations)

    @app.route('/delete/<int:id>')
    def delete(id):
        employee = Employee.query.get_or_404(id)
        try:
            db.session.delete(employee)
            db.session.commit()
            flash('Сотрудник успешно удален!', 'success')
        except Exception as e:
            flash(f"Ошибка удаления сотрудника: {str(e)}", 'danger')
        return redirect(url_for('index'))

    @app.route('/export_excel_xlsx')
    def export_excel_xlsx():
        employees = Employee.query.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"

        headers = [
            'ID', 'ФИО', 'Дата рождения', 'Должность', 'По приказу № 721', 'Состояние', 'Примечание',
            'ВЛК дата', 'ВЛК диагноз', 'КМО дата', 'КМО диагноз', 'УМО дата', 'УМО диагноз',
            'КМО2 дата', 'КМО2 диагноз'
        ]
        ws.append(headers)

        for emp in employees:
            vlk_date = vlk_diagnosis = kmo_date = kmo_diagnosis = umo_date = umo_diagnosis = kmo2_date = kmo2_diagnosis = ''
            for exam in emp.examinations:
                if exam.exam_type == 'ВЛК':
                    vlk_date = exam.exam_date.strftime('%Y-%m-%d')
                    vlk_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'КМО':
                    kmo_date = exam.exam_date.strftime('%Y-%m-%d')
                    kmo_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'УМО':
                    umo_date = exam.exam_date.strftime('%Y-%m-%d')
                    umo_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'КМО2':
                    kmo2_date = exam.exam_date.strftime('%Y-%m-%d')
                    kmo2_diagnosis = exam.diagnosis or ''

            row = [
                emp.id, emp.fio,
                emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '',
                emp.position, emp.order_no, emp.preflight_condition, emp.note or '',
                vlk_date, vlk_diagnosis, kmo_date, kmo_diagnosis, umo_date, umo_diagnosis,
                kmo2_date, kmo2_diagnosis
            ]
            ws.append(row)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='employees.xlsx'
        )

    @app.route('/export_excel_xls')
    def export_excel_xls():
        employees = Employee.query.all()
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Сотрудники')

        headers = [
            'ID', 'ФИО', 'Дата рождения', 'Должность', 'По приказу № 721', 'Состояние', 'Примечание',
            'ВЛК дата', 'ВЛК диагноз', 'КМО дата', 'КМО диагноз', 'УМО дата', 'УМО диагноз',
            'КМО2 дата', 'КМО2 диагноз'
        ]
        for col, header in enumerate(headers):
            ws.write(0, col, header)

        row_num = 1
        for emp in employees:
            vlk_date = vlk_diagnosis = kmo_date = kmo_diagnosis = umo_date = umo_diagnosis = kmo2_date = kmo2_diagnosis = ''
            for exam in emp.examinations:
                if exam.exam_type == 'ВЛК':
                    vlk_date = exam.exam_date.strftime('%Y-%m-%d')
                    vlk_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'КМО':
                    kmo_date = exam.exam_date.strftime('%Y-%m-%d')
                    kmo_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'УМО':
                    umo_date = exam.exam_date.strftime('%Y-%m-%d')
                    umo_diagnosis = exam.diagnosis or ''
                elif exam.exam_type == 'КМО2':
                    kmo2_date = exam.exam_date.strftime('%Y-%m-%d')
                    kmo2_diagnosis = exam.diagnosis or ''

            row = [
                emp.id, emp.fio,
                emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '',
                emp.position, emp.order_no, emp.preflight_condition, emp.note or '',
                vlk_date, vlk_diagnosis, kmo_date, kmo_diagnosis, umo_date, umo_diagnosis,
                kmo2_date, kmo2_diagnosis
            ]
            for col, value in enumerate(row):
                ws.write(row_num, col, value)
            row_num += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.ms-excel',
            as_attachment=True,
            download_name='employees.xls'
        )

    @app.route('/import_excel', methods=['GET', 'POST'])
    def import_excel():
        if request.method == 'POST':
            if 'file' not in request.files:
                flash('Файл не выбран!', 'danger')
                return redirect(url_for('import_excel'))
            file = request.files['file']
            if file.filename == '':
                flash('Файл не выбран!', 'danger')
                return redirect(url_for('import_excel'))
            if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
                try:
                    if file.filename.endswith('.xlsx'):
                        wb = load_workbook(file)
                        ws = wb.active
                        rows = list(ws.rows)
                        headers = [cell.value for cell in rows[0]]
                        data_rows = rows[1:]
                    else:
                        wb = xlrd.open_workbook(file_contents=file.read())
                        ws = wb.sheet_by_index(0)
                        headers = ws.row_values(0)
                        data_rows = [ws.row_values(i) for i in range(1, ws.nrows)]

                    skipped_count = 0
                    for row in data_rows:
                        if file.filename.endswith('.xlsx'):
                            data = [cell.value if cell.value is not None else '' for cell in row]
                        else:
                            data = [str(cell) if cell else '' for cell in row]

                        fio = data[1]
                        existing_employee = Employee.query.filter_by(fio=fio).first()
                        if existing_employee:
                            skipped_count += 1
                            continue

                        birth_date_str = data[2]
                        if birth_date_str:
                            try:
                                birth_date = datetime.strptime(str(birth_date_str), '%Y-%m-%d')
                            except ValueError:
                                flash(f"Неверный формат даты рождения для {fio}: {birth_date_str}", 'danger')
                                return redirect(url_for('import_excel'))
                        else:
                            birth_date = None

                        employee = Employee(
                            fio=fio,
                            birth_date=birth_date,
                            position=data[3],
                            order_no=data[4],
                            preflight_condition=data[5] if data[5] in PREFLIGHT_CONDITIONS else 'Допущен',
                            note=data[6] if data[6] else None
                        )
                        db.session.add(employee)
                        db.session.flush()

                        for i, exam_type in enumerate(['ВЛК', 'КМО', 'УМО', 'КМО2']):
                            date_idx = 7 + i * 2
                            diag_idx = 8 + i * 2
                            exam_date_str = data[date_idx]
                            if exam_date_str:
                                try:
                                    exam_date = datetime.strptime(str(exam_date_str), '%Y-%m-%d')
                                    exam = Examination(
                                        employee_id=employee.id,
                                        exam_type=exam_type,
                                        exam_date=exam_date,
                                        diagnosis=data[diag_idx] if data[diag_idx] else None
                                    )
                                    db.session.add(exam)
                                except ValueError:
                                    flash(f"Неверный формат даты осмотра {exam_type} для {fio}: {exam_date_str}",
                                          'danger')
                                    return redirect(url_for('import_excel'))

                    db.session.commit()
                    flash(f'Сотрудники успешно импортированы из Excel! Пропущено дубликатов: {skipped_count}',
                          'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f"Ошибка импорта Excel: {str(e)}", 'danger')
                return redirect(url_for('index'))
        return render_template('import_excel.html')